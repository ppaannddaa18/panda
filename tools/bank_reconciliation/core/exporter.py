"""结果导出器"""

import os
from datetime import datetime
from typing import List, Dict, Any
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows

import sys
from pathlib import Path as SysPath
sys.path.insert(0, str(SysPath(__file__).parent.parent.parent))
from bank_reconciliation.models import MatchResult, MatchStatus


class ResultExporter:
    """结果导出器"""

    # 样式定义
    HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    HEADER_FONT = Font(bold=True, color="FFFFFF")
    MATCHED_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    UNMATCHED_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    BORDER = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )

    def __init__(self, output_dir: str = "output"):
        """
        初始化导出器

        Args:
            output_dir: 输出目录
        """
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(parents=True, exist_ok=True)

    def export_all(self, results: List[MatchResult]) -> Dict[str, str]:
        """
        导出所有结果

        Returns:
            输出文件路径字典
        """
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")

        return {
            "match_result": self.export_match_result(results, timestamp),
            "balance_sheet": self.export_balance_sheet(results, timestamp),
            "unmatched_details": self.export_unmatched_details(results, timestamp),
            "statistics": self.export_statistics(results, timestamp)
        }

    def export_match_result(self, results: List[MatchResult], timestamp: str = None) -> str:
        """导出对账结果 Excel"""
        if timestamp is None:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")

        file_path = self.output_dir / f"对账结果_{timestamp}.xlsx"

        # 构建数据
        data = []
        for result in results:
            row = {
                "匹配状态": result.status_display,
                "银行日期": result.bank_txn.date.strftime("%Y-%m-%d") if result.bank_txn else "-",
                "银行摘要": result.bank_txn.summary if result.bank_txn else "-",
                "银行借方": float(result.bank_txn.amount) if result.bank_txn and result.bank_txn.direction == "debit" else "",
                "银行贷方": float(result.bank_txn.amount) if result.bank_txn and result.bank_txn.direction == "credit" else "",
                "账务日期": result.account_txn.date.strftime("%Y-%m-%d") if result.account_txn else "-",
                "账务摘要": result.account_txn.summary if result.account_txn else "-",
                "账务借方": float(result.account_txn.amount) if result.account_txn and result.account_txn.direction == "debit" else "",
                "账务贷方": float(result.account_txn.amount) if result.account_txn and result.account_txn.direction == "credit" else "",
                "凭证号": result.account_txn.voucher_no if result.account_txn else "-",
                "匹配说明": result.match_reason
            }
            data.append(row)

        df = pd.DataFrame(data)

        # 创建工作簿
        wb = Workbook()
        ws = wb.active
        ws.title = "对账结果"

        # 写入数据
        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                cell = ws.cell(row=r_idx, column=c_idx, value=value)

                # 应用边框
                cell.border = self.BORDER

                # 表头样式
                if r_idx == 1:
                    cell.fill = self.HEADER_FILL
                    cell.font = self.HEADER_FONT
                    cell.alignment = Alignment(horizontal="center")
                else:
                    # 条件格式
                    status = data[r_idx - 2]["匹配状态"]
                    if status in ["已达账项", "手动匹配", "拆分匹配"]:
                        cell.fill = self.MATCHED_FILL
                    else:
                        cell.fill = self.UNMATCHED_FILL

        # 调整列宽
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            ws.column_dimensions[column_letter].width = min(max_length + 2, 30)

        wb.save(file_path)
        return str(file_path)

    def export_balance_sheet(self, results: List[MatchResult], timestamp: str = None) -> str:
        """导出余额调节表"""
        if timestamp is None:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")

        file_path = self.output_dir / f"余额调节表_{timestamp}.xlsx"

        wb = Workbook()
        ws = wb.active
        ws.title = "余额调节表"

        # 标题
        ws.merge_cells("A1:F1")
        ws["A1"] = "银行存款余额调节表"
        ws["A1"].font = Font(bold=True, size=14)
        ws["A1"].alignment = Alignment(horizontal="center")

        # 分类统计
        categories = {
            "企业已收银行未收": [],
            "企业已付银行未付": [],
            "银行已收企业未收": [],
            "银行已付企业未付": []
        }

        for result in results:
            if result.status == MatchStatus.ACCOUNT_UNMATCHED:
                if result.account_txn.direction == "credit":
                    categories["企业已收银行未收"].append(result.account_txn)
                else:
                    categories["企业已付银行未付"].append(result.account_txn)
            elif result.status == MatchStatus.BANK_UNMATCHED:
                if result.bank_txn.direction == "credit":
                    categories["银行已收企业未收"].append(result.bank_txn)
                else:
                    categories["银行已付企业未付"].append(result.bank_txn)

        row = 3
        for category, txns in categories.items():
            ws.cell(row=row, column=1, value=category).font = Font(bold=True)
            row += 1

            for txn in txns:
                ws.cell(row=row, column=1, value=txn.date.strftime("%Y-%m-%d"))
                ws.cell(row=row, column=2, value=txn.summary)
                ws.cell(row=row, column=3, value=float(txn.amount))
                row += 1

            # 小计
            total = sum(float(t.amount) for t in txns)
            ws.cell(row=row, column=1, value="小计").font = Font(bold=True)
            ws.cell(row=row, column=3, value=total).font = Font(bold=True)
            row += 2

        # 应用边框
        for row_cells in ws.iter_rows(min_row=3, max_row=row, min_col=1, max_col=3):
            for cell in row_cells:
                cell.border = self.BORDER

        wb.save(file_path)
        return str(file_path)

    def export_unmatched_details(self, results: List[MatchResult], timestamp: str = None) -> str:
        """导出未达账项明细"""
        if timestamp is None:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")

        file_path = self.output_dir / f"未达账项明细_{timestamp}.xlsx"

        unmatched = [r for r in results if not r.is_matched]

        data = []
        for result in unmatched:
            if result.bank_txn:
                data.append({
                    "类型": "银行未达",
                    "日期": result.bank_txn.date.strftime("%Y-%m-%d"),
                    "摘要": result.bank_txn.summary,
                    "金额": float(result.bank_txn.amount),
                    "方向": "借方" if result.bank_txn.direction == "debit" else "贷方"
                })
            if result.account_txn:
                data.append({
                    "类型": "企业未达",
                    "日期": result.account_txn.date.strftime("%Y-%m-%d"),
                    "摘要": result.account_txn.summary,
                    "金额": float(result.account_txn.amount),
                    "方向": "借方" if result.account_txn.direction == "debit" else "贷方"
                })

        df = pd.DataFrame(data)
        df.to_excel(file_path, index=False)

        return str(file_path)

    def export_statistics(self, results: List[MatchResult], timestamp: str = None) -> str:
        """导出对账统计报告"""
        if timestamp is None:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")

        file_path = self.output_dir / f"对账统计报告_{timestamp}.xlsx"

        # 统计数据
        total = len(results)
        matched = sum(1 for r in results if r.is_matched)
        bank_unmatched = sum(1 for r in results if r.status == MatchStatus.BANK_UNMATCHED)
        account_unmatched = sum(1 for r in results if r.status == MatchStatus.ACCOUNT_UNMATCHED)

        data = {
            "指标": [
                "总记录数",
                "已匹配数",
                "银行未达数",
                "企业未达数",
                "匹配率"
            ],
            "数值": [
                total,
                matched,
                bank_unmatched,
                account_unmatched,
                f"{matched / total * 100:.1f}%" if total > 0 else "0%"
            ]
        }

        df = pd.DataFrame(data)
        df.to_excel(file_path, index=False)

        return str(file_path)
