"""结果导出器"""

from datetime import datetime
from typing import List, Dict
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side, PatternFill, Alignment

import sys
sys.path.insert(0, str(Path(__file__).parent.parent.parent))
from excel_merger.models import MergeResult, SourceRecord


class ResultExporter:
    """结果导出器"""

    # 样式定义
    HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    HEADER_FONT = Font(bold=True, color="FFFFFF")
    TOTAL_FILL = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    BORDER = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )

    def __init__(self, output_dir: str = "output"):
        """
        初始化导出器

        Args:
            output_dir: 输出目录
        """
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(parents=True, exist_ok=True)

    def export_all(self, result: MergeResult) -> Dict[str, str]:
        """
        导出所有结果

        Args:
            result: 合并结果

        Returns:
            输出文件路径字典
        """
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")

        return {
            "pivot_table": self.export_pivot_table(result, timestamp),
            "unmatched": self.export_unmatched_report(result, timestamp)
        }

    def export_pivot_table(self, result: MergeResult, timestamp: str = None) -> str:
        """
        导出透视表

        Args:
            result: 合并结果
            timestamp: 时间戳

        Returns:
            文件路径
        """
        if timestamp is None:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")

        file_path = self.output_dir / f"合并汇总表_{timestamp}.xlsx"

        # 构建透视表数据
        records = result.records
        if not records:
            # 创建空文件
            wb = Workbook()
            ws = wb.active
            ws.title = "汇总表"
            ws["A1"] = "无数据"
            wb.save(file_path)
            return str(file_path)

        # 创建 DataFrame
        data = []
        for r in records:
            data.append({
                "科目": r.standardized_name,
                "公司": r.company,
                "月份": r.month,
                "金额": float(r.amount)
            })

        df = pd.DataFrame(data)

        # 创建透视表：科目 × (公司-月份)
        pivot = df.pivot_table(
            index="科目",
            columns=["公司", "月份"],
            values="金额",
            aggfunc="sum",
            fill_value=0
        )

        # 创建工作簿
        wb = Workbook()
        ws = wb.active
        ws.title = "汇总表"

        # 写入表头
        ws["A1"] = "科目"

        # 获取所有公司和月份组合
        companies = sorted(df["公司"].unique())
        months = sorted(df["月份"].unique())

        # 写入多级表头
        col_idx = 2
        company_start_cols = {}  # 记录每个公司的起始列

        for company in companies:
            company_start_cols[company] = col_idx
            for month in months:
                ws.cell(row=1, column=col_idx, value=company)
                ws.cell(row=2, column=col_idx, value=f"{month}月")
                col_idx += 1

        # 合并公司名称单元格
        for company in companies:
            start_col = company_start_cols[company]
            end_col = start_col + len(months) - 1
            if len(months) > 1:
                ws.merge_cells(start_row=1, start_column=start_col, end_row=1, end_column=end_col)

        # 应用表头样式
        for col in range(1, col_idx):
            cell = ws.cell(row=1, column=col)
            cell.fill = self.HEADER_FILL
            cell.font = self.HEADER_FONT
            cell.alignment = Alignment(horizontal="center")
            cell.border = self.BORDER

            cell2 = ws.cell(row=2, column=col)
            cell2.fill = self.HEADER_FILL
            cell2.font = self.HEADER_FONT
            cell2.alignment = Alignment(horizontal="center")
            cell2.border = self.BORDER

        # 写入数据
        accounts = sorted(df["科目"].unique())
        for row_idx, account in enumerate(accounts, start=3):
            ws.cell(row=row_idx, column=1, value=account).border = self.BORDER

            for col_offset, company in enumerate(companies):
                for month_offset, month in enumerate(months):
                    col = 2 + col_offset * len(months) + month_offset
                    try:
                        value = pivot.loc[account, (company, month)]
                        cell = ws.cell(row=row_idx, column=col, value=value)
                        cell.number_format = '#,##0.00'
                    except KeyError:
                        cell = ws.cell(row=row_idx, column=col, value=0)
                    cell.border = self.BORDER

        # 添加合计行
        total_row = 3 + len(accounts)
        ws.cell(row=total_row, column=1, value="合计").font = Font(bold=True)
        ws.cell(row=total_row, column=1).fill = self.TOTAL_FILL
        ws.cell(row=total_row, column=1).border = self.BORDER

        for col in range(2, col_idx):
            # 计算列合计
            total = 0
            for row in range(3, total_row):
                val = ws.cell(row=row, column=col).value
                if val:
                    total += val
            cell = ws.cell(row=total_row, column=col, value=total)
            cell.number_format = '#,##0.00'
            cell.font = Font(bold=True)
            cell.fill = self.TOTAL_FILL
            cell.border = self.BORDER

        # 调整列宽
        ws.column_dimensions["A"].width = 20
        for col in range(2, col_idx):
            ws.column_dimensions[ws.cell(row=2, column=col).column_letter].width = 12

        wb.save(file_path)
        return str(file_path)

    def export_unmatched_report(self, result: MergeResult, timestamp: str = None) -> str:
        """
        导出未匹配项报告

        Args:
            result: 合并结果
            timestamp: 时间戳

        Returns:
            文件路径
        """
        if timestamp is None:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")

        file_path = self.output_dir / f"未匹配项报告_{timestamp}.xlsx"

        if not result.unmatched_accounts:
            # 创建空报告
            wb = Workbook()
            ws = wb.active
            ws.title = "未匹配项"
            ws["A1"] = "所有科目均已匹配"
            wb.save(file_path)
            return str(file_path)

        # 统计未匹配科目出现次数
        account_counts: Dict[str, int] = {}
        account_sources: Dict[str, List[str]] = {}

        for record in result.records:
            if record.account_name in result.unmatched_accounts:
                if record.account_name not in account_counts:
                    account_counts[record.account_name] = 0
                    account_sources[record.account_name] = []
                account_counts[record.account_name] += 1
                source = f"{record.company} - {record.source_file}"
                if source not in account_sources[record.account_name]:
                    account_sources[record.account_name].append(source)

        # 创建报告
        wb = Workbook()
        ws = wb.active
        ws.title = "未匹配项"

        # 表头
        headers = ["原始科目名称", "出现次数", "来源文件"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = self.HEADER_FILL
            cell.font = self.HEADER_FONT
            cell.border = self.BORDER

        # 数据
        for row, (account, count) in enumerate(sorted(account_counts.items()), 2):
            ws.cell(row=row, column=1, value=account).border = self.BORDER
            ws.cell(row=row, column=2, value=count).border = self.BORDER
            ws.cell(row=row, column=3, value="; ".join(account_sources[account][:3])).border = self.BORDER

        # 调整列宽
        ws.column_dimensions["A"].width = 30
        ws.column_dimensions["B"].width = 12
        ws.column_dimensions["C"].width = 50

        wb.save(file_path)
        return str(file_path)
