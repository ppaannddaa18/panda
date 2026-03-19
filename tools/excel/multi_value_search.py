"""
Excel 多值搜索工具

支持在 Excel 文件中搜索多个值，并显示匹配结果。
功能特性：
- 支持拖拽文件加载
- 支持 2-3 个值的组合搜索
- 支持包含/精确匹配模式
- 支持搜索单元格批注
- 结果预览和导出功能
"""

import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import openpyxl
from tkinterdnd2 import DND_FILES, TkinterDnD
import threading
import os
from typing import Dict, List, Any, Optional, Tuple


class ExcelSearchApp:
    """Excel多值搜索工具主应用类"""

    def __init__(self, root: TkinterDnD.Tk) -> None:
        self.root = root
        self.root.title("多值搜索 V2.0")
        self.root.geometry("800x850")
        self.root.minsize(800, 850)

        # 初始化样式和数据
        self._setup_styles()
        self._init_data()
        self._setup_ui()

    def _setup_styles(self) -> None:
        """设置应用样式"""
        style = ttk.Style()
        try:
            style.theme_use('clam')
        except tk.TclError:
            pass

        # 统一样式配置
        style.configure('Title.TLabel', font=('Arial', 12, 'bold'))
        style.configure('Heading.TLabel', font=('Arial', 10, 'bold'))
        style.configure('Success.TLabel', foreground='green', font=('Arial', 9))
        style.configure('Error.TLabel', foreground='red', font=('Arial', 9))
        style.configure('Info.TLabel', foreground='blue', font=('Arial', 9))
        style.configure('SearchValue.TLabelframe', font=('Arial', 9, 'bold'), foreground='#2c5aa0')
        style.configure('SearchMode.TLabelframe', font=('Arial', 9, 'bold'), foreground='#2c5aa0')
        style.configure('Options.TLabelframe', font=('Arial', 9, 'bold'), foreground='#2c5aa0')
        style.configure('Action.TButton', font=('Arial', 10, 'bold'))
        style.configure('Toggle.TButton', font=('Arial', 9))

        self.root.configure(bg='#f0f0f0')

    def _init_data(self) -> None:
        """初始化数据存储"""
        self.workbook_data: Dict[str, List[List[Dict[str, Any]]]] = {}
        self.file_path: str = ""
        self.current_preview_item: Optional[Dict[str, Any]] = None

    def _setup_ui(self) -> None:
        """设置用户界面"""
        # 主容器
        self.main_container = ttk.Frame(self.root)
        self.main_container.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)

        # 创建各区域
        self._create_title_section()
        self._create_file_section()
        self._create_search_section()
        self._create_progress_section()
        self._create_result_preview_section()
        self._create_status_section()

    def _create_title_section(self) -> None:
        """创建标题区域"""
        title_frame = ttk.Frame(self.main_container)
        title_frame.pack(fill=tk.X, pady=(0, 15))

        ttk.Label(title_frame, text="多值搜索", style='Title.TLabel').pack(side=tk.LEFT)
        ttk.Label(title_frame, text="V2.0", font=('Arial', 8), foreground='gray').pack(side=tk.RIGHT)
        ttk.Separator(self.main_container, orient='horizontal').pack(fill=tk.X, pady=(0, 10))

    def _create_file_section(self) -> None:
        """创建文件选择区域"""
        file_frame = ttk.LabelFrame(self.main_container, text=" 文件选择 ", padding=15)
        file_frame.pack(fill=tk.X, pady=(0, 10))

        # 拖拽区域
        self.drag_frame = tk.Frame(file_frame, bg='#e8f4fd', relief='ridge', bd=2, height=80)
        self.drag_frame.pack(fill=tk.X, pady=(0, 10))
        self.drag_frame.pack_propagate(False)

        inner_frame = tk.Frame(self.drag_frame, bg='#e8f4fd', relief='groove', bd=1)
        inner_frame.pack(fill=tk.BOTH, expand=True, padx=3, pady=3)

        self.drag_label = tk.Label(
            inner_frame,
            text="拖拽Excel文件到此处 或 点击选择文件\n支持 .xlsx 和 .xls 格式",
            bg='#e8f4fd', font=('Arial', 10), foreground='#2c5aa0', cursor='hand2'
        )
        self.drag_label.pack(expand=True)

        # 启用拖拽
        for widget in [self.drag_frame, inner_frame, self.drag_label]:
            widget.drop_target_register(DND_FILES)
            widget.dnd_bind('<<Drop>>', self.on_file_drop)
        self.drag_label.bind("<Button-1>", self.select_file)

        # 文件信息
        info_frame = ttk.Frame(file_frame)
        info_frame.pack(fill=tk.X)
        ttk.Label(info_frame, text="当前文件:", font=('Arial', 9)).pack(side=tk.LEFT)
        self.file_path_var = tk.StringVar(value="未选择文件")
        ttk.Label(info_frame, textvariable=self.file_path_var, style='Info.TLabel').pack(side=tk.LEFT, padx=(5, 0))
        ttk.Button(info_frame, text="浏览文件", command=self.select_file, style='Action.TButton').pack(side=tk.RIGHT)

    def _create_search_section(self) -> None:
        """创建搜索设置区域"""
        search_frame = ttk.LabelFrame(self.main_container, text=" 搜索设置 ", padding=15)
        search_frame.pack(fill=tk.X, pady=(0, 10))

        main_container = ttk.Frame(search_frame)
        main_container.pack(fill=tk.X)

        # 左侧：搜索值输入
        values_container = ttk.Frame(main_container)
        values_container.pack(side=tk.LEFT, fill=tk.Y, padx=(0, 10))

        # 创建搜索值输入框
        self._create_search_value_inputs(values_container)

        # 右侧：选项设置
        options_container = ttk.Frame(main_container)
        options_container.pack(side=tk.RIGHT, fill=tk.BOTH, expand=True)

        # 搜索模式
        self._create_search_modes(options_container)

        # 其他选项
        self._create_other_options(options_container)

        # 搜索按钮
        ttk.Button(search_frame, text="开始搜索", command=self.start_search, style='Action.TButton').pack(
            pady=(10, 0), ipadx=20, ipady=5)

    def _create_search_value_inputs(self, parent: ttk.Frame) -> None:
        """创建搜索值输入框"""
        self.value_vars = [tk.StringVar() for _ in range(3)]
        self.value_frames = []

        for i in range(3):
            frame = ttk.LabelFrame(parent, text=f" 搜索值 {i + 1} ", style='SearchValue.TLabelframe')
            frame.pack(fill=tk.X, pady=(0, 8))

            entry = ttk.Entry(frame, textvariable=self.value_vars[i], font=('Arial', 10), width=25)
            entry.pack(fill=tk.X, padx=5, pady=5)

            self.value_frames.append(frame)
            if i == 2:  # 第三个值默认隐藏
                frame.pack_forget()

    def _create_search_modes(self, parent: ttk.Frame) -> None:
        """创建搜索模式设置"""
        mode_frame = ttk.LabelFrame(parent, text=" 搜索模式 ", style='SearchMode.TLabelframe')
        mode_frame.pack(fill=tk.X, pady=(0, 10))

        self.search_modes = [tk.StringVar(value="contain") for _ in range(3)]
        self.mode_frames = []

        for i in range(3):
            frame = ttk.Frame(mode_frame)
            frame.pack(fill=tk.X, pady=2)

            ttk.Label(frame, text=f"值{i + 1}:", font=('Arial', 9)).pack(side=tk.LEFT)
            ttk.Radiobutton(frame, text="包含", variable=self.search_modes[i], value="contain").pack(side=tk.LEFT, padx=(10, 5))
            ttk.Radiobutton(frame, text="精确", variable=self.search_modes[i], value="exact").pack(side=tk.LEFT, padx=(0, 5))

            self.mode_frames.append(frame)
            if i == 2:  # 第三个模式默认隐藏
                frame.pack_forget()

    def _create_other_options(self, parent: ttk.Frame) -> None:
        """创建其他选项"""
        options_frame = ttk.LabelFrame(parent, text=" 其他选项 ", style='Options.TLabelframe')
        options_frame.pack(fill=tk.X, pady=(0, 10))

        self.search_comments = tk.BooleanVar(value=True)
        ttk.Checkbutton(options_frame, text="搜索单元格批注", variable=self.search_comments).pack(anchor=tk.W, padx=5, pady=5)

        self.enable_value3 = tk.BooleanVar(value=False)
        self.toggle_btn = ttk.Button(
            options_frame, text="+ 启用第三个搜索值", command=self.toggle_value3, style='Toggle.TButton'
        )
        self.toggle_btn.pack(fill=tk.X, padx=5, pady=(0, 5))

    def _create_progress_section(self) -> None:
        """创建进度显示区域"""
        progress_frame = ttk.Frame(self.main_container)
        progress_frame.pack(fill=tk.X, pady=(0, 10))

        self.progress = ttk.Progressbar(progress_frame, mode='indeterminate', length=400)
        self.progress.pack(side=tk.LEFT, fill=tk.X, expand=True)

        self.progress_label = ttk.Label(progress_frame, text="", font=('Arial', 9))
        self.progress_label.pack(side=tk.RIGHT, padx=(10, 0))

    def _create_result_preview_section(self) -> None:
        """创建结果和预览区域"""
        self.paned_window = tk.PanedWindow(self.main_container, orient=tk.HORIZONTAL, sashwidth=8, sashrelief=tk.RAISED)
        self.paned_window.pack(fill=tk.BOTH, expand=True, pady=(0, 10))

        # 结果区域
        result_frame = ttk.LabelFrame(self.paned_window, text=" 搜索结果 ", padding=10)
        self._setup_result_area(result_frame)
        self.paned_window.add(result_frame, minsize=400)

        # 预览区域
        self.preview_frame = ttk.LabelFrame(self.paned_window, text=" 内容预览 ", padding=10)
        self._setup_preview_area(self.preview_frame)
        self.paned_window.add(self.preview_frame, minsize=200)

    def _setup_result_area(self, parent: ttk.Frame) -> None:
        """设置结果区域"""
        # 结果统计
        info_frame = ttk.Frame(parent)
        info_frame.pack(fill=tk.X, pady=(0, 10))

        self.result_count_var = tk.StringVar(value="等待搜索...")
        ttk.Label(info_frame, textvariable=self.result_count_var, font=('Arial', 9, 'bold')).pack(side=tk.LEFT)

        self.export_btn = ttk.Button(info_frame, text="导出结果", command=self.export_results, state='disabled')
        self.export_btn.pack(side=tk.RIGHT)

        # Treeview容器
        tree_container = ttk.Frame(parent)
        tree_container.pack(fill=tk.BOTH, expand=True)

        # 初始Treeview（6列）
        columns = ("工作表", "行号", "值1位置", "值1内容", "值2位置", "值2内容")
        self.result_tree = ttk.Treeview(tree_container, columns=columns, show="headings", height=15)

        self._configure_treeview(columns, False)

        # 滚动条
        self._add_scrollbars(tree_container, self.result_tree)

        # 事件绑定
        self.result_tree.bind("<ButtonRelease-1>", self.on_result_click)
        self.result_tree.bind("<Button-3>", self.show_context_menu)

    def _setup_preview_area(self, parent: ttk.Frame) -> None:
        """设置预览区域"""
        preview_container = ttk.Frame(parent)
        preview_container.pack(fill=tk.BOTH, expand=True)

        self.preview_text = tk.Text(preview_container, wrap=tk.WORD, font=('Consolas', 10), bg='white', relief='sunken', bd=2)
        self._add_scrollbars(preview_container, self.preview_text)

        # 初始提示
        self.preview_text.insert(tk.END, "点击搜索结果中的任意内容单元格\n查看完整内容预览")
        self.preview_text.config(state=tk.DISABLED)

    def _configure_treeview(self, columns: Tuple[str, ...], enable_value3: bool) -> None:
        """配置Treeview列"""
        column_widths = {
            "工作表": 70,
            "行号": 50,
            "值1位置": 50,
            "值1内容": 100,
            "值2位置": 50,
            "值2内容": 100,
            "值3位置": 50,
            "值3内容": 100
        }

        for col in columns:
            self.result_tree.heading(col, text=col, anchor=tk.W)
            self.result_tree.column(col, width=column_widths.get(col, 100), anchor=tk.W)

    def _add_scrollbars(self, parent: ttk.Frame, widget: tk.Widget) -> None:
        """为组件添加滚动条"""
        v_scrollbar = ttk.Scrollbar(parent, orient="vertical", command=widget.yview)
        h_scrollbar = ttk.Scrollbar(parent, orient="horizontal", command=widget.xview)

        if isinstance(widget, ttk.Treeview):
            widget.configure(yscrollcommand=v_scrollbar.set, xscrollcommand=h_scrollbar.set)
        elif isinstance(widget, tk.Text):
            widget.configure(yscrollcommand=v_scrollbar.set, xscrollcommand=h_scrollbar.set)

        widget.grid(row=0, column=0, sticky="nsew")
        v_scrollbar.grid(row=0, column=1, sticky="ns")
        h_scrollbar.grid(row=1, column=0, sticky="ew")

        parent.grid_rowconfigure(0, weight=1)
        parent.grid_columnconfigure(0, weight=1)

    def _create_status_section(self) -> None:
        """创建状态栏"""
        status_frame = ttk.Frame(self.main_container, relief='sunken', padding=5)
        status_frame.pack(fill=tk.X)

        self.status_var = tk.StringVar(value="请选择Excel文件开始搜索")
        ttk.Label(status_frame, textvariable=self.status_var, font=('Arial', 9)).pack(side=tk.LEFT)

        self.time_label = ttk.Label(status_frame, text="", font=('Arial', 8), foreground='gray')
        self.time_label.pack(side=tk.RIGHT)
        self.update_time()

    def update_time(self) -> None:
        """更新时间显示"""
        import datetime
        current_time = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        self.time_label.config(text=current_time)
        self.root.after(1000, self.update_time)

    def toggle_value3(self) -> None:
        """切换第三个值的显示/隐藏"""
        if self.enable_value3.get():
            # 隐藏第三个值
            self.value_frames[2].pack_forget()
            self.mode_frames[2].pack_forget()
            self.toggle_btn.config(text="+ 启用第三个搜索值")
            self.enable_value3.set(False)
        else:
            # 显示第三个值
            self.value_frames[2].pack(fill=tk.X, pady=(0, 8))
            self.mode_frames[2].pack(fill=tk.X, pady=2)
            self.toggle_btn.config(text="- 禁用第三个搜索值")
            self.enable_value3.set(True)

    def on_file_drop(self, event: Any) -> None:
        """处理文件拖拽"""
        try:
            files = self.root.tk.splitlist(event.data)
            if files:
                file_path = files[0]
                if file_path.lower().endswith(('.xlsx', '.xls')):
                    self.load_excel_file(file_path)
                else:
                    messagebox.showerror("文件格式错误", "请选择Excel文件(.xlsx或.xls)")
        except Exception as e:
            messagebox.showerror("错误", f"处理拖拽文件时出错: {str(e)}")

    def select_file(self, event: Optional[tk.Event] = None) -> None:
        """选择文件"""
        file_path = filedialog.askopenfilename(
            title="选择Excel文件",
            filetypes=[
                ("Excel files", "*.xlsx *.xls"),
                ("Excel 2007-365", "*.xlsx"),
                ("Excel 97-2003", "*.xls"),
                ("All files", "*.*")
            ]
        )
        if file_path:
            self.load_excel_file(file_path)

    def load_excel_file(self, file_path: str) -> None:
        """加载Excel文件"""
        try:
            self.status_var.set("正在加载文件...")
            self.progress_label.config(text="加载中...")
            self.progress.start()

            threading.Thread(target=self._load_file_thread, args=(file_path,), daemon=True).start()
        except Exception as e:
            self.progress.stop()
            self.progress_label.config(text="")
            messagebox.showerror("错误", f"加载文件失败: {str(e)}")
            self.status_var.set("加载文件失败")

    def _load_file_thread(self, file_path: str) -> None:
        """在线程中加载文件"""
        try:
            self.file_path = file_path
            self.workbook_data = {}

            workbook = openpyxl.load_workbook(file_path, data_only=False)

            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                sheet_data = []

                max_row = sheet.max_row
                max_col = sheet.max_column

                if max_row is None or max_col is None:
                    continue

                for row in range(1, max_row + 1):
                    row_data = []
                    for col in range(1, max_col + 1):
                        cell = sheet.cell(row=row, column=col)
                        cell_info = {
                            'value': cell.value,
                            'comment': cell.comment.text if cell.comment else None,
                            'coordinate': cell.coordinate
                        }
                        row_data.append(cell_info)
                    sheet_data.append(row_data)

                self.workbook_data[sheet_name] = sheet_data

            self.root.after(0, self._update_ui_after_load)
        except Exception as e:
            self.root.after(0, lambda: self._handle_load_error(str(e)))

    def _update_ui_after_load(self) -> None:
        """加载完成后更新UI"""
        self.progress.stop()
        self.progress_label.config(text="")

        filename = os.path.basename(self.file_path)
        self.file_path_var.set(filename)
        self.status_var.set(f"文件加载完成，共{len(self.workbook_data)}个工作表")

        # 更新拖拽区域显示
        self.drag_label.config(
            text=f"文件已加载: {filename}\n点击可重新选择文件",
            bg='#e8f5e8', foreground='#2d5016'
        )
        self.drag_frame.config(bg='#e8f5e8')

        # 更新内部框架
        for child in self.drag_frame.winfo_children():
            if isinstance(child, tk.Frame):
                child.config(bg='#e8f5e8')

    def _handle_load_error(self, error_msg: str) -> None:
        """处理加载错误"""
        self.progress.stop()
        self.progress_label.config(text="")
        messagebox.showerror("错误", f"加载文件失败: {error_msg}")
        self.status_var.set("加载文件失败")

    def start_search(self) -> None:
        """开始搜索"""
        if not self.workbook_data:
            messagebox.showwarning("警告", "请先加载Excel文件")
            return

        values = [var.get().strip() for var in self.value_vars]
        value1, value2, value3 = values[0], values[1], values[2] if self.enable_value3.get() else ""

        if not value1 or not value2:
            messagebox.showwarning("警告", "请输入至少两个搜索值")
            return

        if value1 == value2 or (value3 and (value1 == value3 or value2 == value3)):
            messagebox.showwarning("警告", "搜索值不能相同")
            return

        # 清空结果和预览
        for item in self.result_tree.get_children():
            self.result_tree.delete(item)

        self.preview_text.config(state=tk.NORMAL)
        self.preview_text.delete(1.0, tk.END)
        self.preview_text.insert(tk.END, "点击搜索结果中的任意内容单元格\n查看完整内容预览")
        self.preview_text.config(state=tk.DISABLED)
        self.current_preview_item = None

        self.result_count_var.set("搜索中...")
        self.status_var.set("正在搜索...")
        self.progress_label.config(text="搜索中...")
        self.progress.start()
        self.export_btn.config(state='disabled')

        threading.Thread(target=self._search_thread, args=(value1, value2, value3), daemon=True).start()

    def _search_thread(self, value1: str, value2: str, value3: str) -> None:
        """在线程中执行搜索"""
        try:
            results = []
            search_modes = [mode.get() for mode in self.search_modes]
            search_comments = self.search_comments.get()

            for sheet_name, sheet_data in self.workbook_data.items():
                for row_idx, row_data in enumerate(sheet_data, 1):
                    found_positions = []

                    for cell_info in row_data:
                        cell_value = cell_info['value']
                        cell_comment = cell_info['comment']
                        coordinate = cell_info['coordinate']

                        # 检查所有值
                        for i, (search_val, mode) in enumerate(zip([value1, value2, value3], search_modes)):
                            if not search_val:
                                continue

                            if self._check_match(cell_value, search_val, mode):
                                found_positions.append({
                                    'coordinate': coordinate,
                                    'found_value': search_val,
                                    'original_value': cell_value,
                                    'type': 'cell',
                                    'search_target': f'value{i + 1}'
                                })

                            if search_comments and cell_comment and self._check_match(cell_comment, search_val, mode):
                                found_positions.append({
                                    'coordinate': coordinate,
                                    'found_value': search_val,
                                    'original_value': cell_comment,
                                    'type': 'comment',
                                    'search_target': f'value{i + 1}'
                                })

                    if self._contains_all_values(found_positions, value1, value2, value3):
                        positions = {}
                        for pos in found_positions:
                            target = pos['search_target']
                            if target not in positions:
                                positions[target] = pos

                        if 'value1' in positions and 'value2' in positions:
                            result_item = {
                                'sheet': sheet_name,
                                'row': row_idx,
                                'value1_pos': positions['value1'],
                                'value2_pos': positions['value2'],
                            }
                            if value3 and 'value3' in positions:
                                result_item['value3_pos'] = positions['value3']
                            results.append(result_item)

            self.root.after(0, lambda: self._update_search_results(results))
        except Exception as e:
            self.root.after(0, lambda: self._handle_search_error(str(e)))

    @staticmethod
    def _check_match(cell_value: Any, search_value: str, search_mode: str) -> bool:
        """检查是否匹配"""
        if cell_value is None:
            return False

        cell_str = str(cell_value).strip()
        search_str = str(search_value).strip()

        if search_mode == "exact":
            return cell_str == search_str
        else:
            return search_str in cell_str

    @staticmethod
    def _contains_all_values(found_positions: List[Dict[str, Any]], value1: str, value2: str, value3: str) -> bool:
        """检查是否包含所有需要的值"""
        targets = {pos['search_target'] for pos in found_positions}
        required = {'value1', 'value2'}
        if value3:
            required.add('value3')
        return required.issubset(targets)

    def _update_search_results(self, results: List[Dict[str, Any]]) -> None:
        """更新搜索结果"""
        self.progress.stop()
        self.progress_label.config(text="")

        # 重新配置Treeview
        enable_value3 = self.enable_value3.get()
        columns = ("工作表", "行号", "值1位置", "值1内容", "值2位置", "值2内容", "值3位置",
                   "值3内容") if enable_value3 else ("工作表", "行号", "值1位置", "值1内容", "值2位置", "值2内容")

        self.result_tree.configure(columns=columns, show="headings")
        self._configure_treeview(columns, enable_value3)

        for result in results:
            values = [
                result['sheet'],
                result['row'],
                f"{result['value1_pos']['coordinate']}({result['value1_pos']['type']})",
                str(result['value1_pos']['original_value']),
                f"{result['value2_pos']['coordinate']}({result['value2_pos']['type']})",
                str(result['value2_pos']['original_value'])
            ]

            if enable_value3 and 'value3_pos' in result:
                values.extend([
                    f"{result['value3_pos']['coordinate']}({result['value3_pos']['type']})",
                    str(result['value3_pos']['original_value'])
                ])

            self.result_tree.insert("", "end", values=values)

        # 更新统计
        if results:
            self.result_count_var.set(f"找到 {len(results)} 个匹配结果")
            self.status_var.set(f"搜索完成，找到 {len(results)} 个结果")
            self.export_btn.config(state='normal')
        else:
            self.result_count_var.set("未找到匹配结果")
            self.status_var.set("未找到匹配的行")
            messagebox.showinfo("搜索结果", "未找到匹配的行")

    def _handle_search_error(self, error_msg: str) -> None:
        """处理搜索错误"""
        self.progress.stop()
        self.progress_label.config(text="")
        messagebox.showerror("错误", f"搜索过程中出错: {error_msg}")
        self.status_var.set("搜索失败")
        self.result_count_var.set("搜索失败")

    def on_result_click(self, event: tk.Event) -> None:
        """点击结果行的处理"""
        region = self.result_tree.identify("region", event.x, event.y)
        if region == "cell":
            column = self.result_tree.identify_column(event.x)
            item_id = self.result_tree.identify_row(event.y)

            if item_id:
                item = self.result_tree.item(item_id)
                values = item['values']
                self.current_preview_item = item

                # 检查点击的是内容列
                if column in ['#4', '#6', '#8']:
                    col_index = int(column[1:]) - 1
                    if col_index < len(values):
                        content = str(values[col_index])
                        self.preview_text.config(state=tk.NORMAL)
                        self.preview_text.delete(1.0, tk.END)
                        self.preview_text.insert(tk.END, content)
                        self.preview_text.config(state=tk.DISABLED)

    def show_context_menu(self, event: tk.Event) -> None:
        """显示右键菜单"""
        item_id = self.result_tree.identify_row(event.y)
        if item_id:
            self.result_tree.selection_set(item_id)
            self.context_menu = tk.Menu(self.root, tearoff=0)
            self.context_menu.add_command(label="复制值1位置坐标", command=self.copy_value1_position)
            self.context_menu.add_command(label="复制值2位置坐标", command=self.copy_value2_position)
            if self.enable_value3.get():
                self.context_menu.add_command(label="复制值3位置坐标", command=self.copy_value3_position)
            self.context_menu.add_separator()
            self.context_menu.add_command(label="复制行信息", command=self.copy_row_info)
            self.context_menu.add_separator()
            self.context_menu.add_command(label="删除此行", command=self.delete_selected_row)

            try:
                self.context_menu.tk_popup(event.x_root, event.y_root)
            finally:
                self.context_menu.grab_release()

    def copy_value1_position(self) -> None:
        """复制值1位置坐标"""
        self._copy_position(2, "值1位置坐标")

    def copy_value2_position(self) -> None:
        """复制值2位置坐标"""
        self._copy_position(4, "值2位置坐标")

    def copy_value3_position(self) -> None:
        """复制值3位置坐标"""
        self._copy_position(6, "值3位置坐标")

    def _copy_position(self, index: int, label: str) -> None:
        """复制位置坐标的通用方法"""
        selection = self.result_tree.selection()
        if selection:
            item = self.result_tree.item(selection[0])
            values = item['values']
            if len(values) > index:
                coordinate = values[index].split('(')[0]
                self.root.clipboard_clear()
                self.root.clipboard_append(coordinate)
                self.status_var.set(f"{label} {coordinate} 已复制到剪贴板")

    def copy_row_info(self) -> None:
        """复制行信息"""
        selection = self.result_tree.selection()
        if selection:
            item = self.result_tree.item(selection[0])
            values = item['values']
            info = "\t".join(str(v) for v in values)
            self.root.clipboard_clear()
            self.root.clipboard_append(info)
            self.status_var.set("行信息已复制到剪贴板")

    def delete_selected_row(self) -> None:
        """删除选中行"""
        selection = self.result_tree.selection()
        if selection:
            self.result_tree.delete(selection[0])
            remaining = len(self.result_tree.get_children())
            self.result_count_var.set(f"剩余 {remaining} 个匹配结果")

    def export_results(self) -> None:
        """导出搜索结果"""
        if not self.result_tree.get_children():
            messagebox.showwarning("警告", "没有可导出的结果")
            return

        file_path = filedialog.asksaveasfilename(
            title="保存搜索结果",
            defaultextension=".csv",
            filetypes=[("CSV files", "*.csv"), ("Text files", "*.txt"), ("All files", "*.*")]
        )

        if file_path:
            try:
                with open(file_path, 'w', encoding='utf-8-sig') as f:
                    # 写入标题
                    headers = self.result_tree['columns']
                    f.write(','.join(headers) + '\n')

                    # 写入数据
                    for item in self.result_tree.get_children():
                        values = self.result_tree.item(item)['values']
                        line = ','.join(f'"{str(v)}"' for v in values)
                        f.write(line + '\n')

                messagebox.showinfo("成功", f"结果已导出到: {file_path}")
                self.status_var.set(f"结果已导出到: {os.path.basename(file_path)}")
            except Exception as e:
                messagebox.showerror("错误", f"导出失败: {str(e)}")


def main() -> None:
    """主函数"""
    try:
        root = TkinterDnD.Tk()
        ExcelSearchApp(root)

        # 居中窗口
        root.update_idletasks()
        width = root.winfo_width()
        height = root.winfo_height()
        x = (root.winfo_screenwidth() // 2) - (width // 2)
        y = (root.winfo_screenheight() // 2) - (height // 2)
        root.geometry(f'{width}x{height}+{x}+{y}')

        root.mainloop()
    except ImportError:
        messagebox.showerror("错误", "请安装 tkinterdnd2 库: pip install tkinterdnd2")
    except Exception as e:
        messagebox.showerror("错误", f"程序启动失败: {str(e)}")


if __name__ == "__main__":
    main()